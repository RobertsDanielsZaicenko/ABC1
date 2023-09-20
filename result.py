from openpyxl import load_workbook
wb = load_workbook("test1.xlsx")
ws=wb.active
max_row=ws.max_row
total=0
for i in range(2,max_row+1):
    hours=ws['C'+str(i)].value
    rate=ws['B'+str(i)].value
    if (type(hours)!=str and type(rate)!=str):
         salary=(hours*rate)
         if salary>3000:
              total=total+1
print(total)
wb.save('ŗesult.xlsx')
wb.close()

